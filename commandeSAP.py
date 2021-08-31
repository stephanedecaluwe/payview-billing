# -*- coding: utf-8 -*-
# conversion synthèse ADV .xlsx => fichier de commande SAP .csv
# nov. 2020 
# librairies openpyxl remplacant xlrd deprecated

import openpyxl
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import sys
import os

racine=r"c:\users\sdecaluwe\Desktop\factupayview\code\Generated"
path_file="C:/Users/sdecaluwe/Desktop/factupayview/code/Generated/20210706_15h08_facturation_juin_2021/20210706_15h08_juin_2021_facturationGlobale.xlsx"

#path_file = sys.argv[1]
#mois = sys.argv[2]
#annee = sys.argv[3]
mois = "6"
annee = "2021"

nom_fichier="C:/Users/sdecaluwe/Desktop/factupayview/code/Generated/synthese ADV payview.xlsx"

PO_number ="facturation mensuelle "+mois+"."+annee

periode_long = mois+"."+annee

repertoire_syntheses=racine+"\\"

path_lut = racine+"..\input\LUT.xlsx"

path_concat = racine+"\concatentation.csv"

overfee_exonerated = ["ACCOR", "ADIDAS FRANCE", "AEM SOFTS", "AGAPES", "AMESYS CONSEIL", "AUCHAN HYPERMARCHE", "B & B HOTELS", "BOULANGER", "BUFFALO GRILL", "CABESTO", "CABINET OTP", "CARREFOUR BELGIQUE", "CARREFOUR FRANCE", "CARREFOUR ITALIE", "CGR", "CONDUENT", "COURIR FRANCE", "DATACAR", "ELIOR", "ERAM", "ETAM LINGERIE", "EUROTUNNEL", "FFT", "GO SPORT", "GRAND FRAIS (FUJITSU)", "JULES", "HEURTAUX SAS", "IRIS INFORMATIQUE", "JCDECAUX FRANCE", "KBANE", "KFC FRANCE SAS", "LA FERME DU SART", "LE BON MARCHE", "LE FURET DU NORD", "Leroy Merlin Italy", "MCDONALD'S FRANCE", "METRO FRANCE", "NORAUTO FRANCE", "PARAGON ID", "PARKARE FRANCE", "PETIT-BATEAU", "PRINTEMPS", "PROMOD", "RATP", "SOCIETE AIR FRANCE", "SODEXO", "STIME", "SUPERMARCHES MATCH", "TAPIS SAINT MACLOU", "TISSEO", "TOSHIBA", "VIVATICKET", "WASHTEC FRANCE", "WAYCOM INTERNATIONAL", "AVT", "AVT Grossiste", "HM TELECOM", "HM TELECOM Grossiste", "IPSF", "NEOSYSTEMS_IPSF", "IPSF DIRECT", "LAVANCE EMIC", "CAFES MERLING", "SELECTA", "THE BOX OFFICE COMPANY", "LM CONTROL", "LM CONTROL Grossiste", "MONEY30", "MONEY30 Grossiste", "ORANGE"]


orga_tss = "8101"
orga_ms = "7140"
division="8120" #delivery plant
poste_tarif_nul="ZSEF" # colonne i 
poste_tarif_nonnul="ZSEV"
liste_exclusion = [  ]
liste_exclusion_client = [ ]
dictionnaire_substitution={}
tarif0=[ ]







for i in os.listdir(racine):
   if i.endswith("synthese ADV payivew.xlsx"):
      os.remove(racine+"/"+i)


###
wb_input=Workbook()
wb_input = load_workbook(filename=path_file)
ws_input = wb_input["Facturation"]
for i in range(1,10):
   row=[]
   for j in range(1,10):
      row.append(ws_input.cell(i,j).value)
      

wb_output = Workbook(write_only=True)
ws_output = wb_output.create_sheet('Synthèse ADV')


#for row in rows_sim_passport:
#    if len(row) > 6:
#        if row[3]!="Contract":
#            ws_output.append(row)
#            #print(row)
#wb_output_sim_passport.save("resultat_sim_passport.xlsx")
###


if len(mois)==1:
    mois = "0"+mois


codeSAP=[0]*15
row_title=["code SAP","client","identifiants","Activité","Opération","Logiciel","Article de prestation","Libellé","Quantité","Prix","Montant","Entre","et","BU"]
ws_output.append(row_title)

for row in range(2, ws_input.max_row+1):
   for column in range(1, ws_input.max_column+1):
       cell = str(ws_input.cell(row,column).value)
       if column==1:
          client = cell
       if column==3:
          bu=cell
       if column==4:
          codeSAP=cell
       if column>=6:
          quantity = cell
          if quantity=="None":
             quantity="0"
          libelle = ws_input.cell(1,column).value
          if libelle == "PAS_SIM_OVERFEE" and client in overfee_exonerated:
             quantity = "0"
          if codeSAP!="" and codeSAP!="None" and int(quantity)!=0: 
             if mois in ["01","03","05","07","08","10","12"]:
                 row_to_add=[codeSAP, client, "", "service", "service", "",libelle,"", int(quantity), "", "","01."+str(mois)+"."+str(annee),"31."+str(mois)+"."+str(annee),bu]
             else:
                row_to_add=[codeSAP, client, "", "service", "service", "",libelle,"", int(quantity), "", "","01."+str(mois)+"."+str(annee),"30."+str(mois)+"."+str(annee),bu]
             if mois=="02":
                row_to_add=[codeSAP, client, "", "service", "service", "",libelle,"", int(quantity), "", "","01."+str(mois)+"."+str(annee),"28."+str(mois)+"."+str(annee),bu]
             ws_output.append(row_to_add)
wb_output.save(nom_fichier)


# effacement de SAP/*.csv
for i in os.listdir(repertoire_syntheses):
   if i.endswith(".csv"):
      os.remove(repertoire_syntheses+i)
      
fichier_total = open(path_concat,'a')   

# pour tous les fichiers du repertoire de sortie de tem connect type synthèse ADV*.xlsx
for i in os.listdir(repertoire_syntheses):
    # si le fichier s'appelle synthese adv*.xlsx
    if i.startswith("synthese ADV") and (i.endswith(".xlsx") or i.endswith(".XLSX")):
       print("traitemenent de ........ "+i)
       wb = load_workbook(filename = repertoire_syntheses+i)
       # aller dans le premier onglet (numero 0)
       ws = wb.worksheets[0]
       # la date de commande SAP (PO) est prise dans le champ "à" de la synthèse ADV.xlsx (ex : du 01/07/2020 à 30/09/2020)
       PO_date=str(ws.cell(2,13).value)
       # le code client SAP est pris dans la colonne A (indice 0)
       print(ws)
       print(ws.cell(2,1).value)
       code_client= str(ws.cell(2,1).value)
       code_client_precedent=42
       #organisation = dico_entite[code_client]
       business_unit=str(ws.cell(2,14).value)
       if business_unit == "TSS":
          organisation=orga_tss
       else:
          organisation=orga_ms
       # le nom du client, dans la colonne B (indice 1)
       nom_client=str(ws.cell(2,2).value)
       # assemblage des chaines selon le format final : nom fichier, ORG, HEADER.
       nom_fichier=repertoire_syntheses+nom_client+" "+PO_date.replace("/","-")+".csv"
       fichier  = open(nom_fichier,'a')
       for row in range(ws.max_row):
          # exclusion des noms de champs
          
          if ws.cell(row+1,1).value != "" and ws.cell(row+1,1).value !="code SAP" and ws.cell(row+1,1).value !="Code SAP" and ws.cell(row+1,1).value !="code SAP client":
             # code SAP produit, quantite;
             code_client= str(ws.cell(row+1,1).value)
             #print(code_client)
             if code_client!=code_client_precedent and code_client not in liste_exclusion_client:
                # un nouveau client est traité : ORG, HEADER et TEXTH sont mis en en-tête
                print(code_client)
                business_unit = str(ws.cell(row+1,14).value)
                if business_unit == "TSS":
                   organisation=orga_tss
                else:
                   organisation= orga_ms
                fichier.write("ORG;ZSCE;"+str(organisation)+";Z1;SE;;;;\n")
                fichier.write("HEADER;"+PO_number+";"+PO_date+";;"+code_client+";"+code_client+";;;\n")
                fichier.write("TEXTH;0011;FR;Facturation PAYVIEW;;;;;\n")
                fichier.write("TEXTH;0011;FR;Periode :  "+periode_long+";;;;;\n")
                fichier.write("TEXTH;0011;FR;Merci d'adresser votre demande de justificatif a :;;;;;\n")
                fichier.write("TEXTH;0011;FR;adv-telechargements@ingenico.com;;;;;\n")
                fichier_total.write("ORG;ZSCE;"+str(organisation)+";Z1;SE;;;;\n")
                fichier_total.write("HEADER;"+PO_number+";"+PO_date+";;"+code_client+";"+code_client+";;;\n")
                fichier_total.write("TEXTH;0011;FR;Facturation PAYVIEW;;;;;\n")
                fichier_total.write("TEXTH;0011;FR;Periode :  "+periode_long+";;;;;\n")
                fichier_total.write("TEXTH;0011;FR;Merci d'adresser votre demande de justificatif a :;;;;;\n")
                fichier_total.write("TEXTH;0011;FR;adv-telechargements@ingenico.com;;;;;\n")
                code_client_precedent=code_client
             code_produit = str(ws.cell(row+1,7).value)
             if code_produit == "PAS_SIM1_500_N":
                code_produit = "PAS_SIM_1_500_N"
             if code_produit == "PAS_SIM2_500_N":
                code_produit = "PAS_SIM_2_500_N"
             quantite = str(ws.cell(row+1,9).value)

             # exclusion et substitution du code produit le cas echeant    
             # écriture d'une ligne correspondant à un produit, quantité, client.
             # codes SAP exclus : non reconnus par SAP car à créer.
             # Les codes SAP à tarif nul sont traités avec le poste_tarif_nul (ZSEN, ZSEF ...).
             # application d'une liste de substitution pour garantir le bon format des codes SAP (corrigé maintenant).
             if code_produit not in liste_exclusion and code_client not in liste_exclusion_client:
                if code_produit not in dictionnaire_substitution:
                   if code_produit not in tarif0:
                      fichier.write("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nonnul+"\n")
                      fichier_total.write("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nonnul+"\n")
                      print("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nonnul+"\n")
                      # cas standard : tarif_nonnul
                   else:
                      fichier.write("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nul+"\n")
                      fichier_total.write("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nul+"\n")
                      print("ITEM;"+code_produit+";"+quantite+";;EUR;;"+division+";;"+poste_tarif_nul+"\n")
                      #tarif0 non exclu, non substitué
                else:                 
                   fichier.write("ITEM;"+dictionnaire_substitution[code_produit]+";"+quantite+";;EUR;;"+division+";;\n")
                   fichier_total.write("ITEM;"+dictionnaire_substitution[code_produit]+";"+quantite+";;EUR;;"+division+";;\n")
                   print("ITEM;"+dictionnaire_substitution[code_produit]+";"+quantite+";;EUR;;"+division+";;\n")
                   #tarif standard substitué
       # sauvegarde, fermeture
       fichier.close()
       
fichier_total.write("") 
   
fichier_total.close()

print("fichiers écrits")   
 
       
